import urllib3
import time
import re
import random
import certifi
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver

http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=certifi.where())


def main():
    start = 0
    year_low = 2007
    year_high = 2019
    #   This is our year combination
    #   2007 - 2008 => Index 0
    #   2009 - 2010 => Index 1
    #   2011 - 2012 => Index 2
    #   2013 - 2014 => Index 3
    #   2015 - 2016 => Index 4
    #   2017 - 2018 => Index 5
    for i in range(0, 6):
        start = 0
        for j in range(0, 990):
            delay = random.randrange(3, 15)
            time.sleep(delay)
            page_scrape = 'https://scholar.google.ca/scholar?start=' + str(start) +\
                          '&q=barcodinglife.org&hl=en&as_sdt=0,5&as_ylo='+str(year_low)+'&as_yhi=' + str(year_low + 1)
            r = get_request(page_scrape)
            citation_papers = process_request(r)
            length = len(citation_papers)
            if length < 1:
                r = get_request(page_scrape)
                citation_papers = process_request(r)
                length = len(citation_papers)
                if length < 1:
                    print("::::::::::::::::::::DONE: Year Processed::::::::::::::::::::\n")
                    break
            else:
                print("Processing Tab: " + str(start) +
                      "\nYear Range: " + str(year_low) + " - " + str(year_low + 1) + "\n")
                process_data(citation_papers, start, year_low, year_low + 1)
                start = start + 10
        if year_low == year_high:
            exit(0)
        if year_low < year_high:
            year_low = year_low + 2
            print("::::::::::::::::::::Year Batch Changed::::::::::::::::::::\n")
        else:
            print("::::::::::::::::::::DONE: Accessed Last year::::::::::::::::::::\n")


def process_data(data, page_number, year_st, year_ed):
    # load the workbook
    wb = load_workbook(filename="data_cids.xlsx", data_only=True)

    # Select the Sheet to work with
    sheet = wb['Sheet1']
    row = sheet.max_row + 1

    try:
        for i in range(0, len(data)):
            curr_data = BeautifulSoup(str(data[i]), 'html.parser')
            data_attributes = curr_data.find('div').attrs
            sheet.cell(row=row, column=1).value = str(year_st)
            sheet.cell(row=row, column=2).value = str(year_ed)
            sheet.cell(row=row, column=3).value = str(page_number)
            try:
                sheet.cell(row=row, column=4).value = data_attributes['data-cid']
            except Exception as e:
                print(e)
                print("Exception! No DATA_CID found on Tab: " + str(page_number) + " Year Range: " + str(year_st) +
                      " - " + str(year_ed) + " Item number: " + str(i))
                wb.save("data_cids.xlsx")
                pass
            paper_title = ""
            try:
                title = curr_data.find_all('h3', attrs={'class': 'gs_rt'})
                for j in range(0, len(title)):
                    list_data = BeautifulSoup(str(title[j]), 'html.parser')
                    paper_title_dom = list_data.find('a', attrs={'data-clk': re.compile(r".*")})
                    paper_title = paper_title_dom.string
                if paper_title != "":
                    sheet.cell(row=row, column=5).value = paper_title
            except:
                try:
                    title = curr_data.find('h3', attrs={'class': 'gs_rt'})
                    sheet.cell(row=row, column=5).value = title.text
                except Exception as s:
                    print(s)
                    print("Exception! No Title found on Tab: " + str(page_number) + " Year Range: " + str(year_st)
                          + " - " + str(year_ed) + " Item number: " + str(i))
                    pass
                wb.save("data_cids.xlsx")
                pass
            row = row + 1
    except Exception as e:
        print(e)
        print("Exception Occurred at: " + str(page_number) + " Year Range: " + str(year_st) + " - " + str(year_ed))
        wb.save("data_cids.xlsx")
        pass

    wb.save("data_cids.xlsx")


def get_request(page_scrape):
    r = http.request(
        'GET',
        page_scrape,
        headers={
            "authority": "scholar.google.ca",
            "method": "GET",
            "path": "/scholar?hl=en&as_sdt=0%2C5&sciodt=0%2C5&cites=10992335886072847329&scipsc="
                    "&as_ylo=2007&as_yhi=2018",
            "scheme": "https",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
            "accept-encoding": "gzip, deflate, br",
            "accept-language": "en-US,en;q=0.9",
            "cache-control": "max-age=0",
            "cookie": "CONSENT=YES+CA.en+20170806-09-0; OGPC=19005936-2:19006818-1:19006913-1:19006965-1:19007018-1:"
                      "19007661-1:19008374-2:19009193-2:19009353-2:; GSP=IN=baf1eeca98964b5+4161895ca238f336+bd2811c4f6"
                      "02422c+15ea0d06b0148432:LD=en:CF=3:CO=0:LM=1543508372:S=G2ZfOq8IGYndZxVk; SID=0AawxS6wwzZFsqayhb"
                      "coiewO7JqBsIZvFEq_YPALcGNzImsXCGvEVDtOE_DLvvQ4q5dXuA.; HSID=A3yD5AKVy6iI5wXwS; SSID=AsVpykjncgl"
                      "7mU21n; APISID=teB7JLmgXcTa7VKW/AtZ_xNINcKcJykU7U; SAPISID=EJ7-hll4DavzTh-O/AAQTC6mazc-H7Y2fD;"
                      " 1P_JAR=2018-12-11-20; NID=150=Zq2tFsu_6I4i1c8d6tUKOhhFXm7d2MnaY-ShUhT0hjYpAb8RZodXZ47dNk5cuILp"
                      "ULiWS5DrSL374i6Mac9_i_bjyFOslcqX2T7_D9aZLOMg0Pyg-qA2jS4LIKR0cNgLtB4Vo0BkPd4JJ5RO_-sP_R9XMRNmR5X"
                      "3ywsg8-7QO5eRaQMGGiVNhmBXzIsUgRUirz0XT84PMZWKeSi9L0gzpi0vPr9d0d3rNEBcz2WkaG4VFW7RNIpxOZB-3ZMXP-"
                      "5ZEB0JIvxyU-AX-cMTB8XWm0WaiV7l_iXT_xfHnvghqKyNI9iDxUCxo-Z5ZswzGJ-O_S2Y-bQ9ZQqEwVbO1MVmzHT2tqdv"
                      "sT1V",
            "user-agent": "Mozilla/5.0 (Linux; Android 6.0; P027 Build/MRA58L) AppleWebKit/537.36 (KHTML, like Gecko)"
                          " Chrome/57.0.2987.132 Safari/537.36",
            "x-client-data": "CJK2yQEIprbJAQjEtskBCKmdygEIqKPKAQi/p8oBGPmlygE="
        }
    )

    return r


def process_request(r):
    soup = BeautifulSoup(r.data, 'html.parser')
    citation_papers = soup.find_all('div', attrs={'class': 'gs_r gs_or gs_scl'})
    print(soup)
    return citation_papers


def driver():
    driver = webdriver.Chrome('./chromedriver_win32/chromedriver')
    driver.get('http://www.google.com/xhtml')
    time.sleep(5)  # Let the user actually see something!
    search_box = driver.find_element_by_name('q')
    search_box.send_keys('ChromeDriver')
    search_box.submit()
    time.sleep(5)  # Let the user actually see something!
    driver.quit()


if __name__ == '__main__':
    main()
    # driver()